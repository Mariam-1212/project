from flask import Flask, render_template, request, redirect, url_for, session
from models import Room, Customer, Booking, Hotel, Employee
from datetime import datetime, date
import openpyxl, os
import qrcode
import base64
from io import BytesIO

app = Flask(__name__)
app.secret_key = "hotel_secure_key"


# --------------------------------------------------------
# Create Excel file if not exists
# --------------------------------------------------------
if not os.path.exists("data.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "BookingID", "Name", "Email", "Phone",
        "RoomType", "Guests", "CheckIn", "CheckOut",
        "Total", "Status", "Rating"
    ])
    wb.save("data.xlsx")


# --------------------------------------------------------
# Hotel Setup
# --------------------------------------------------------
hotel = Hotel("Nile View Hotel", "Cairo, Egypt", "+20 100 555 7777")

hotel.addRoom(Room(1, "Single Room", 500, "Cozy single room", 1, 5))
hotel.addRoom(Room(2, "Double Room", 800, "Perfect for couples", 2, 3))
hotel.addRoom(Room(3, "Deluxe Suite", 1500, "Nile view luxury suite", 4, 2))

admin = Employee(1, "Admin", "Manager", "admin", "1234")


# --------------------------------------------------------
# Generate QR Code Invoice
# --------------------------------------------------------
def generate_qr(name, booking_id, room, total, ci, co):
    text = f"""
Booking Invoice
-------------------------
Name: {name}
Booking ID: {booking_id}
Room: {room}
Total: {total} EGP
Check-in: {ci}
Check-out: {co}
"""
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(text)
    qr.make(True)

    img = qr.make_image(fill="black", back_color="white")
    buffer = BytesIO()
    img.save(buffer, format="PNG")
    return base64.b64encode(buffer.getvalue()).decode()


# --------------------------------------------------------
# Home Page
# --------------------------------------------------------
@app.route('/')
def home():
    return render_template("home.html", rooms=hotel.displayAvailableRooms(), hotel=hotel)


# --------------------------------------------------------
# About
# --------------------------------------------------------
@app.route('/about')
def about():
    return render_template("about.html", hotel=hotel)


# --------------------------------------------------------
# Booking Page
# --------------------------------------------------------
@app.route('/book/<int:id>', methods=["GET", "POST"])
def book(id):
    room = next((r for r in hotel.listOfRooms if r.roomNumber == id), None)
    if not room:
        return "Room Not Found", 404

    if request.method == "POST":
        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]
        guests = int(request.form["guests"])
        ci = request.form["check_in"]
        co = request.form["check_out"]

        checkIn = datetime.strptime(ci, "%Y-%m-%d").date()
        checkOut = datetime.strptime(co, "%Y-%m-%d").date()

        if guests > room.maxPeople:
            return render_template("booking.html", room=room, error="Too many guests!")

        if checkOut <= checkIn:
            return render_template("booking.html", room=room, error="Invalid date range.")

        if not room.checkAvailability():
            return render_template("booking.html", room=room, error="Room is fully booked!")

        booking_id = datetime.now().strftime("%Y%m%d%H%M%S")
        customer = Customer(booking_id, name, email, phone)
        booking = Booking(booking_id, customer, room, checkIn, checkOut)
        total = booking.calculateTotalAmount()

        room.bookRoom()

        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        ws.append([booking_id, name, email, phone, room.roomType,
                   guests, ci, co, total, "Pending Payment", None])
        wb.save("data.xlsx")

        return redirect(f"/payment/{booking_id}")

    return render_template("booking.html", room=room)


# --------------------------------------------------------
# Payment Page
# --------------------------------------------------------
@app.route('/payment/<booking_id>', methods=["GET", "POST"])
def payment(booking_id):
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

    booking = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == booking_id:
            booking = row
            break

    if not booking:
        return "Booking Not Found", 404

    if request.method == "POST":
        # Update status to CONFIRMED after payment
        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == booking_id:
                ws.cell(row=row, column=10).value = "Confirmed"
                wb.save("data.xlsx")
                break

        return redirect(f"/success/{booking_id}")

    return render_template("payment.html", booking=booking)


# --------------------------------------------------------
# Success Page
# --------------------------------------------------------
@app.route('/success/<booking_id>')
def success_page(booking_id):
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

    booking = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == booking_id:
            booking = row
            break

    name = booking[1]
    room = booking[4]
    total = booking[8]
    ci = booking[6]
    co = booking[7]

    qr = generate_qr(name, booking_id, room, total, ci, co)

    return render_template("success.html",
                           name=name,
                           booking_id=booking_id,
                           room=room,
                           total=total,
                           check_in=ci,
                           check_out=co,
                           qr=qr)


# --------------------------------------------------------
# Rating Page
# --------------------------------------------------------
@app.route('/rate/<booking_id>', methods=["GET", "POST"])
def rate(booking_id):
    if request.method == "POST":
        stars = int(request.form["stars"])

        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            if str(ws.cell(row=row, column=1).value) == booking_id:
                ws.cell(row=row, column=11).value = stars
                room_type = ws.cell(row=row, column=5).value
                hotel.getRoomByType(room_type).addRating(stars)
                wb.save("data.xlsx")
                break

        return redirect("/")

    return render_template("rate.html", booking_id=booking_id)


# --------------------------------------------------------
# Admin Login
# --------------------------------------------------------
@app.route('/login', methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        u = request.form["username"]
        p = request.form["password"]

        if admin.login(u, p):
            session["admin"] = True
            return redirect("/admin")
        else:
            error = "Wrong username or password!"

    return render_template("login.html", error=error)


# --------------------------------------------------------
# Logout
# --------------------------------------------------------
@app.route('/logout')
def logout():
    session.pop("admin", None)
    return redirect("/")


# --------------------------------------------------------
# Admin Panel
# --------------------------------------------------------
@app.route('/admin')
def admin_panel():
    if "admin" not in session:
        return redirect("/login")

    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

    bookings = [row for row in ws.iter_rows(min_row=2, values_only=True)]

    return render_template("admin.html", bookings=bookings, hotel=hotel)


# --------------------------------------------------------
# Update Booking Status
# --------------------------------------------------------
@app.route('/update_status/<booking_id>/<status>', methods=["POST"])
def update_status(booking_id, status):
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == booking_id:
            ws.cell(row=row, column=10).value = status
            wb.save("data.xlsx")
            break

    return redirect("/admin")


# --------------------------------------------------------
# DELETE Booking
# --------------------------------------------------------
@app.route('/delete/<booking_id>', methods=["POST"])
def delete_booking(booking_id):
    if "admin" not in session:
        return redirect("/login")

    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active

    for row in range(2, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value) == booking_id:
            ws.delete_rows(row)
            wb.save("data.xlsx")
            break

    return redirect("/admin")


# --------------------------------------------------------
# RUN APP
# --------------------------------------------------------
if __name__ == "__main__":
    app.run(debug=True)
